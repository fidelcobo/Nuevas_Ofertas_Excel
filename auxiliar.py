import openpyxl
import csv
import os
from PyQt5 import QtWidgets
from constantes import FIRST_ROW
from listas import list_manufacturer
# from datetime import datetime
import locale

locale.setlocale(locale.LC_ALL, 'FR')


def fill_offer(lista_articulos, carpeta, bid, only_maint, moneda, instance):

    """
    :param lista_arrticulos:
    :param carpeta:
    :param instance:
    :return: No devuelve valores, sino que compone los csv de la oferta
    Este procedimiento clasifica los artículos por fabricante y compone los ficheros csv de salida
    """
    if only_maint:  # Sólo mantenimientos en un fichero único
        hacer_oferta_csv('Mantenimiento', lista_articulos, carpeta, bid, only_maint, moneda, instance )
        return

    for fabr in list_manufacturer:
        lista_fabr =  [p
                       for p in lista_articulos
                       if p.manufacturer.lower() == fabr.lower()]
        hacer_oferta_csv(fabr, lista_fabr, carpeta, bid, only_maint, moneda, instance)


def hacer_oferta_csv(fabr:str, lista:list, directorio, bid:str, only_maint, moneda, instance):

    if lista:
        curr_row = 2
        num_fila = 10
        actual_dir = os.path.dirname(__file__)
        filen = os.path.join(actual_dir, 'plantilla.xlsx')  # Este fichero no se toca. Hace de plantilla
        # sheet_name = 'bom'
        libro = openpyxl.load_workbook(filen)
        hoja = libro.get_active_sheet()

        for items in lista:
            if not only_maint:
                fila = str(curr_row)
                hoja['A' + fila] = num_fila
                hoja['C' + fila] = fabr
                hoja['H' + fila] = 1
                hoja['L' + fila] = 'EA'
                hoja['M' + fila] = moneda
                hoja['N' + fila] = 'Fixed'
                hoja['F' + fila] = items.code
                hoja['G' + fila] = items.descripcion_prod
                hoja['K' + fila] = items.unit
                if items.list_price:
                    hoja['O' + fila] = locale.format('%10.2f', items.list_price)
                else:
                    hoja['O' + fila] = 0
                hoja['Q' + fila] = locale.format('%10.2f', items.coste_prod)
                hoja['P' + fila] = locale.format('%10.2f', items.venta_prod)
                hoja['AF' + fila] = items.tech
                if items.maintenance:  # El ítem tiene mantenimiento. Añadir las dos lineas correspondientes
                    fila = str(curr_row + 1)
                    fila_sig = str(curr_row + 2)
            else:
                fila = str(curr_row)
                fila_sig = str(curr_row + 1)

            if items.maintenance:
                hoja['A' + fila] = num_fila + 1
                hoja['A' + fila_sig] = num_fila + 2
                hoja['B' + fila_sig] = num_fila + 1
                hoja['H' + fila] = 2
                hoja['H' + fila_sig] = 20
                hoja['C' + fila] = 'Dimension Data'
                hoja['C' + fila_sig] = items.manufacturer
                hoja['E' + fila] = items.unit
                hoja['E' + fila_sig] = items.unit
                hoja['F' + fila] = items.sku_uptime
                hoja['F' + fila_sig] = items.backout_name
                hoja['G' + fila] = items.descr_uptime
                hoja['G' + fila_sig] = items.code
                hoja['I' + fila] = items.code
                hoja['I' + fila_sig] = 0
                hoja['J' + fila] = items.manufacturer
                hoja['J' + fila_sig] = ''
                hoja['K' + fila] = 1
                hoja['K' + fila_sig] = 1
                hoja['L' + fila] = 'EA'
                hoja['L' + fila_sig] = 'EA'
                hoja['M' + fila] = moneda
                hoja['M' + fila_sig] = moneda
                hoja['N' + fila] = 'Fixed'
                hoja['N' + fila_sig] = 'Fixed'
                hoja['O' + fila] = locale.format('%10.2f',items.list_price_back)
                hoja['O' + fila_sig] = locale.format('%10.2f',items.list_price_back)
                hoja['P' + fila] = locale.format('%10.2f', float(items.venta_mant * items.durac/12))
                hoja['P' + fila_sig] = locale.format('%10.2f', items.list_price_back)
                hoja['Q' + fila] = locale.format('%10.2f', float(items.cost_unit_manten * items.durac/12))
                hoja['Q' + fila_sig] = locale.format('%10.2f', float(items.coste_unit_back * items.durac/12))
                fecha_init = '{}/{}/{}'.format(items.init_date.day, items.init_date.month,
                                               items.init_date.year)
                fecha_init_limpia = '{}{}{}'.format(str(items.init_date.year),
                                                    str(items.init_date.month).zfill(2),
                                                    str(items.init_date.day).zfill(2))
                fecha_fin = '{}/{}/{}'.format(items.end_date.day, items.end_date.month,
                                              items.end_date.year)
                hoja['X' + fila] = fecha_init
                hoja['X' + fila_sig] = fecha_init
                hoja['Y' + fila] = fecha_fin
                hoja['Y' + fila_sig] = fecha_fin
                hoja['AA' + fila] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(items.durac) +
                                     '#InvoiceInterval=Yearly#InvoiceMode=anticipated')
                hoja['AA' + fila_sig] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(items.durac) +
                                         '#InvoiceInterval=Yearly#InvoiceMode=anticipated')

                hoja['AF' + fila] = items.tech
                hoja['AF' + fila_sig] = items.tech
                hoja['AG' + fila] = items.sn
                hoja['AG' + fila_sig] = items.sn
                curr_row += 2

            if not only_maint:  # Si hay productos, saltamos 3 filas en curr_row
                curr_row += 1
            num_fila += 10

        nombre_fichero_salida_exc = str(bid) + '_'+ fabr + '.xlsx'
        nombre_fichero_salida_csv = str(bid) + '_'+ fabr +  '.csv'

        fichero_salida_excel = os.path.join(directorio, nombre_fichero_salida_exc)
        fichero_salida_csv = os.path.join(directorio, nombre_fichero_salida_csv)

        todo_ok = False

        while not todo_ok:
            try:
                libro.save(fichero_salida_excel)
                todo_ok = True
            except PermissionError:
                QtWidgets.QMessageBox.warning(instance, 'Procesado de oferta',
                                              'Fichero Excel de destino {} ya abierto \n'
                                              'por favor ciérrelo y pulse Aceptar'.format(fichero_salida_excel))

        csv_from_excel(fichero_salida_excel, fichero_salida_csv, instance)
        os.remove(fichero_salida_excel)  # Quitamos el fichero auxiliar Excel


def csv_from_excel(entrada, salida, instance):

        # VARIANTE CON xlrd. Pone los números de línea en float
        # print(entrada)
        # with xlrd.open_workbook(entrada) as wb:
        #     sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
        #     with open(salida, 'w', newline='') as f:
        #         c = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        #         for r in range(sh.nrows):
        #             c.writerow(sh.row_values(r))
        ok = False

        while not ok:
            try:
                wb = openpyxl.load_workbook(entrada)
                sh = wb.get_active_sheet()  # or wb.sheet_by_name('name_of_the_sheet_here')

                with open(salida, 'w', newline='') as f:
                    c = csv.writer(f, dialect='excel', delimiter=';')
                    for r in sh.rows:
                        c.writerow([cell.value for cell in r])
                ok = True

            except PermissionError:
                QtWidgets.QMessageBox.warning(instance, 'Procesado de oferta',
                                              'Fichero csv de destino {} ya abierto \n'
                                              'por favor ciérrelo y pulse Aceptar'.format(salida))


def busca_columnas(sheet: object, lista_busca: list, fila_busca: str) -> object:

    columnas = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z', 'AA','AB','AC','AD','AE']

    result_busca = []

    for items in lista_busca:
        for col in columnas:
            cell = col + fila_busca
            if sheet[cell].value == items:
                result_busca.append(col)
                break
    if len(result_busca) == len(lista_busca):  # Se han encontrado todos los campos relevantes
        ok = True
    else:
        ok = False

    return ok, result_busca


def get_ultima_fila(hoja, col_code):
    fila = FIRST_ROW
    fin = False
    while not fin:
        cell = col_code + str(fila)
        if not hoja[cell].value:
            fin = True
        else:
            fila += 1
            fin = False
    return (fila-1)

# AQUÍ VAN LOS PROCEDIMIENTOS QUE SE HAN RETIRADO Y SUSTITUIDO POR OTROS MÁS EFICIENTES
# def clasificar_articulos(lista_articulos, instance):
#     # Este procedimiento toma la lista de artículos y los reparte en listas particularizadas para cada fabricante
#
#     lista_checkpoint = [p
#                         for p in lista_articulos
#                         if p.manufacturer == 'Checkpoint']
#     instance.lista_checkpoint = lista_checkpoint
#
#     lista_fortinet = [p
#                       for p in lista_articulos
#                       if p.manufacturer == 'Fortinet']
#     instance.lista_fortinet = lista_fortinet
#
#     lista_hp = [p
#                 for p in lista_articulos
#                 if p.manufacturer == 'HP']
#     instance.lista_hp = lista_hp
#
#     lista_f5 = [p
#                 for p in lista_articulos
#                 if p.manufacturer == 'F5 Networks']
#     instance.lista_f5 = lista_f5
#
#     lista_cisco = [p
#                    for p in lista_articulos
#                    if p.manufacturer == 'CISCO']
#     instance.lista_cisco = lista_cisco
#
#     lista_alcatel = [p
#                      for p in lista_articulos
#                      if p.manufacturer == 'Alcatel']
#     instance.lista_alcatel = lista_alcatel
#
#     lista_aruba = [p
#                    for p in lista_articulos
#                    if p.manufacturer == 'Aruba']
#     instance.lista_aruba = lista_aruba
#
#     lista_paloalto = [p
#                       for p in lista_articulos
#                       if p.manufacturer == 'Palo Alto Networks']
#     instance.lista_paloalto = lista_paloalto
#
#     lista_juniper = [p
#                      for p in lista_articulos
#                      if p.manufacturer == 'Juniper']
#     instance.lista_juniper = lista_juniper
#
#     lista_bluecoat = [p
#                       for p in lista_articulos
#                       if p.manufacturer == 'Bluecoat']
#     instance.lista_bluecoat = lista_bluecoat
#
#     lista_brocade = [p
#                      for p in lista_articulos
#                      if p.manufacturer == 'Brocade']
#     instance.lista_brocade = lista_brocade


# def hacer_oferta(directorio, instance):
#
#     oferta_productos('Checkpoint', instance.lista_checkpoint, directorio, instance)
#     oferta_productos('Fortinet', instance.lista_fortinet, directorio, instance)
#     oferta_productos('F5', instance.lista_f5, directorio, instance)
#     oferta_productos('HP', instance.lista_hp, directorio, instance)
#     oferta_productos('Aruba', instance.lista_aruba, directorio, instance)
#     oferta_productos('Palo Alto Networks', instance.lista_paloalto, directorio, instance)
#     oferta_productos('Brocade', instance.lista_brocade, directorio, instance)
#     oferta_productos('Juniper', instance.lista_juniper, directorio, instance)
#     oferta_productos('Bluecoat', instance.lista_bluecoat, directorio, instance)
#     oferta_productos('Alcatel', instance.lista_alcatel, directorio, instance)
#     oferta_productos('CISCO', instance.lista_cisco, directorio, instance)



# def hacer_oferta_ms(lista_articulos, directorio, instance):
#
#     # En primer lugar filtramos los registros que incluyen mantenimiento
#
#     lista_ms = [item
#                 for item in lista_articulos
#                 if item.manten == 'Sí']
#
#     excel_aux = pass_to_excel(lista_ms, instance)
#     fichero_excel_out = os.path.join(directorio, 'ms.xlsx')
#     fichero_csv_out = os.path.join(directorio, 'ms.csv')
#
#     if excel_aux:
#
#         excel_aux.save(fichero_excel_out)
#         csv_from_excel(fichero_excel_out, fichero_csv_out)
#         os.remove(fichero_excel_out)


# def pass_to_excel(lista: [], instance):
#
#     actual_dir = os.path.dirname(__file__)
#
#     # Ahora abrimos el fichero Excel auxiliar de plantilla de MS
#     filen = os.path.join(actual_dir, 'plantilla.xlsx')  # Este fichero no se toca. Hace de plantilla
#     sheet_name = 'ms'
#     if not os.path.exists(filen):
#         instance.QtWidgets.QMessageBox.critical(instance, 'Procesado de oferta',
#                                                 'El fichero {} no se encuentra\n'.format(filen))
#         return None
#
#     try:
#         libro = openpyxl.load_workbook(filen)
#         hoja = libro.get_sheet_by_name(sheet_name)
#
#     except PermissionError:
#         instance.QtWidgets.QMessageBox.critical(instance, 'Procesado de oferta',
#                                                'Fichero Excel {} no se puede abrir\n'
#                                                'Compruebe que no está ya abierto'.format(filen))
#
#         return None
#
#     curr_row = 2
#     num_fila = 10
#
#     for reg in lista:
#         fila = str(curr_row)
#         fila_sig = str(curr_row + 1)
#
#         hoja['A' + fila] = num_fila
#         hoja['A' + fila_sig] = num_fila + 1
#         hoja['B' + fila_sig] = num_fila
#         hoja['H' + fila] = 2
#         hoja['H' + fila_sig] = 20
#         hoja['C' + fila] = 'Dimension Data'
#         hoja['C' + fila_sig] = reg.manufacturer
#         hoja['E' + fila] = reg.unit
#         hoja['E' + fila_sig] = reg.unit
#         hoja['F' + fila] = reg.sku_uptime
#         hoja['F' + fila_sig] = reg.backout_name
#         hoja['G' + fila] = reg.descr_uptime
#         hoja['G' + fila_sig] = reg.code
#         hoja['I' + fila] = reg.code
#         hoja['I' + fila_sig] = 0
#         hoja['J' + fila] = reg.manufacturer
#         hoja['J' + fila_sig] = ''
#         hoja['K' + fila] = 1
#         hoja['K' + fila_sig] = 1
#         hoja['L' + fila] = 'EA'
#         hoja['L' + fila_sig] = 'EA'
#         hoja['M' + fila] = 'EUR'
#         hoja['M' + fila_sig] = 'EUR'
#         hoja['N' + fila] = 'Fixed'
#         hoja['N' + fila_sig] = 'Fixed'
#         hoja['O' + fila] = reg.list_price_back
#         hoja['O' + fila_sig] = reg.list_price_back
#         hoja['P' + fila] = locale.format('%10.2f', float(reg.venta_mant))
#         hoja['P' + fila_sig] = locale.format('%10.2f', (1.2 * float(reg.coste_unit_back)))
#         hoja['Q' + fila] = locale.format('%10.2f', float(reg.cost_unit_manten))
#         hoja['Q' + fila_sig] = locale.format('%10.2f', float(reg.coste_unit_back))
#         fecha_init = '{}/{}/{}'.format(reg.init_date.day, reg.init_date.month, reg.init_date.year)
#         fecha_init_limpia = '{}{}{}'.format(str(reg.init_date.year), str(reg.init_date.month).zfill(2),
#                                             str(reg.init_date.day).zfill(2))
#         fecha_fin = '{}/{}/{}'.format(reg.end_date.day, reg.end_date.month, reg.end_date.year)
#         hoja['X' + fila] = fecha_init
#         hoja['X' + fila_sig] = fecha_init
#         hoja['Y' + fila] = fecha_fin
#         hoja['Y' + fila_sig] = fecha_fin
#         hoja['AA' + fila] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(reg.durac) +
#                             '#InvoiceInterval=Yearly#InvoiceMode=anticipated')
#         hoja['AA' + fila_sig] = ('StartDate=' + fecha_init_limpia + '#Duration=' + str(reg.durac) +
#                             '#InvoiceInterval=Yearly#InvoiceMode=anticipated')
#
#         hoja['AF' + fila] = reg.tech
#         hoja['AF' + fila_sig] = reg.tech
#         curr_row += 2
#         num_fila += 3
#
#     return libro